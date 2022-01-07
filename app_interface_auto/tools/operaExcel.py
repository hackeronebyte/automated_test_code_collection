import sys
from tools.file_system_path import file_system_path

sys.path.append(file_system_path())
from openpyxl import *
from openpyxl.styles import *
from tools.get_time import get_now_time_y_m_d
import os


class OperaExcel:
    '''excel 操作'''

    def create_excel():
        '''
        创建日志excle表格，保存日志信息
        :return:
        '''
        # path = "../../case_data_file/" + str(get_now_time_y_m_d()) + "_request_result.xlsx"
        path = file_system_path() + "/case_data_file/" + str(get_now_time_y_m_d()) + "_request_result.xlsx"
        isExists = os.path.exists(path)
        if not isExists:
            # 请求时间	url	header	data	result	code
            data = ['请求时间', 'url', 'header', 'data', 'result', 'code']
            # 在内存创建一个工作簿obj
            wb = Workbook()
            ws = wb.active
            ws.title = u'request_result'
            # 向第一个sheet页写数据
            c = 1
            for line in data:
                ws.cell(row=1, column=c, value=line).font = Font(bold=True, color='000000')
                c += 1
            # 工作簿保存到磁盘
            wb.save(path)
        return path

    def get_file(file):
        case_file = file
        return case_file

    def get_wb(case_file):
        # 打开已有
        wb = load_workbook(OperaExcel.get_file(case_file))
        return wb

    # 加载文件
    def get_excel(case_file):
        wsa = OperaExcel.get_wb(case_file).active
        return wsa

    # 获取最大行
    def get_lines(case_file):
        ws = OperaExcel.get_excel(case_file)
        lines = ws.max_row
        return lines

    # 获取最大列
    def get_colmun(case_file):
        coll = OperaExcel.get_excel(case_file).max_column
        return coll

    # 获取指定单元格数据
    def get_cell(case_file, r, c):
        rows = []
        for row in OperaExcel.get_excel(case_file).iter_rows():
            # print(row)
            rows.append(row)
        data = rows[r][c].value
        # print(data)
        return data

    def save_kp(*args):
        '''
        登录结果保存 login.xlsx
        :param args:
        :return:
        '''
        # True 3 9 10 ../case_data_file/login.xlsx
        # print("save_kp 参数： ", *args)
        wb = load_workbook(OperaExcel.get_file(args[args.__len__() - 1]))
        ws1 = wb.active
        if args[0] == False:
            ws1.cell(row=args[1], column=args[2], value=args[3]).font = Font(bold=True, color='FF0000')  # 预期结果
            ws1.cell(row=args[1], column=args[2] + 1, value="Fail").font = Font(bold=True, color='FF0000')
        else:
            ws1.cell(row=args[1], column=args[2], value=args[3]).font = Font(bold=True, color='000000')
            ws1.cell(row=args[1], column=args[2] + 1, value="Success").font = Font(bold=True, color='000000')

        wb.save(OperaExcel.get_file(args[args.__len__() - 1]))

    def write_cell_kp(*args):

        OperaExcel.save_kp(*args)

    def save_order_result(*args):
        '''
        写入订单结果表 order_run_result.xlsx
        :param args:
        :return:
        '''
        # lines + 1, dict_str["run_time"], data_str, dict_str["cart_totalPrice"],
        #                                      dict_str["order_number"], dict_str["message"], file_name_excel

        # print("save_kp 参数： ", *args)
        wb = load_workbook(OperaExcel.get_file(args[args.__len__() - 1]))
        ws1 = wb.active
        ws1.cell(row=args[0], column=1, value=args[1]).font = Font(bold=False, color='000000')
        ws1.cell(row=args[0], column=2, value=args[2]).font = Font(bold=False, color='000000')
        ws1.cell(row=args[0], column=3, value=args[3]).font = Font(bold=False, color='000000')
        ws1.cell(row=args[0], column=4, value=args[4]).font = Font(bold=False, color='000000')
        ws1.cell(row=args[0], column=5, value=args[5]).font = Font(bold=False, color='000000')
        ws1.cell(row=args[0], column=6, value=args[6]).font = Font(bold=False, color='000000')
        ws1.cell(row=args[0], column=7, value=args[7]).font = Font(bold=False, color='000000')
        ws1.cell(row=args[0], column=8, value=args[8]).font = Font(bold=False, color='000000')
        ws1.cell(row=args[0], column=9, value=args[9]).font = Font(bold=False, color='000000')
        ws1.cell(row=args[0], column=10, value=args[10]).font = Font(bold=False, color='000000')
        ws1.cell(row=args[0], column=11, value=args[11]).font = Font(bold=False, color='000000')
        ws1.cell(row=args[0], column=12, value=args[12]).font = Font(bold=False, color='000000')
        wb.save(OperaExcel.get_file(args[args.__len__() - 1]))

    def request_result(*args):
        '''
        写入请求结果表 request_result.xlsx
        :param args:
        :return:
        '''
        # gettime() lines + 1, url, header, data, result, result["code"], file_path
        wb = load_workbook(OperaExcel.get_file(args[args.__len__() - 1]))
        ws1 = wb.active
        if args[6] == "200":
            ws1.cell(row=args[0], column=1, value=args[1]).font = Font(bold=False, color='000000')
            ws1.cell(row=args[0], column=2, value=args[2]).font = Font(bold=False, color='000000')
            ws1.cell(row=args[0], column=3, value=args[3]).font = Font(bold=False, color='000000')
            ws1.cell(row=args[0], column=4, value=args[4]).font = Font(bold=False, color='000000')
            ws1.cell(row=args[0], column=5, value=args[5]).font = Font(bold=False, color='000000')
            ws1.cell(row=args[0], column=6, value=args[6]).font = Font(bold=False, color='000000')
        else:
            ws1.cell(row=args[0], column=1, value=args[1]).font = Font(bold=True, color='FF0000')
            ws1.cell(row=args[0], column=2, value=args[2]).font = Font(bold=True, color='FF0000')
            ws1.cell(row=args[0], column=3, value=args[3]).font = Font(bold=True, color='FF0000')
            ws1.cell(row=args[0], column=4, value=args[4]).font = Font(bold=True, color='FF0000')
            ws1.cell(row=args[0], column=5, value=args[5]).font = Font(bold=True, color='FF0000')
            ws1.cell(row=args[0], column=6, value=args[6]).font = Font(bold=True, color='FF0000')

        wb.save(OperaExcel.get_file(args[args.__len__() - 1]))


if __name__ == '__main__':
    print(OperaExcel.create_excel())
